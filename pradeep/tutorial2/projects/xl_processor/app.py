
input_file = 'pradeep/tutorial2/projects/xl_processor/transactions.xlsx'
output_file = 'pradeep/tutorial2/projects/xl_processor/corrected_transactions.xlsx'


def process_spreadsheet(input_file, output_file):
    import openpyxl as xl
    from openpyxl.chart import BarChart, Reference

    workbook = xl.load_workbook(input_file)

    sheet = workbook['Sheet']
    cell = sheet['a1']
    cell = sheet.cell(1, 1)

    print(cell.value)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        print(cell.value)
        corrected_value = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_value

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart)

    workbook.save(output_file)
