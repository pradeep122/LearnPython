import openpyxl as xl  # as is used to given an alias to the package
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']  # a is column and 1 is row
    # we can als get the coordinates by using sheet.cell object
    cell = sheet.cell(1, 1)
    print(cell.value)
    print(f'number of rows = {sheet.max_row}')  # to find number of rows

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 4)
        # print(cell.value)
        cell_mrp = sheet.cell(row, 5)
        # print(cell_mrp.value)
        total = cell.value * cell_mrp.value
        total_cell = sheet.cell(row, 6)
        total_cell.value = total

    # values = Reference(sheet, min_row=2, max_row=sheet.max_row,
    #                  min_col=6, max_col=6)

    #chart = BarChart()
    # chart.add_data(values)
    #sheet.add_chart(chart, 'g2')

    wb.save(filename)
