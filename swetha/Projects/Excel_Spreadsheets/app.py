import openpyxl as xl
from openpyxl.chart import BarChart, Reference

input_filename = 'swetha/Projects/Excel_Spreadsheets/transactions.xlsx'
output_filename = 'swetha/Projects/Excel_Spreadsheets/transactions2.xlsx'


def process_workbook(input_filename, output_filename):
    wb = xl.load_workbook(input_filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    print(cell)
    cell = sheet.cell(1, 1)
    print(cell)

    # Iterating to the rows
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2 ')

    wb.save(output_filename)
